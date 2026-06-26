"""
Tools for the resume agent pipeline.
"""

import os
import json
import re
from io import BytesIO
from pathlib import Path
from datetime import datetime

import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
)
from pypdf import PdfReader


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_workbook() -> Path:
    env_path = os.getenv("RESUME_EXCEL_PATH")
    if env_path:
        p = Path(env_path)
        if p.exists():
            return p
        raise FileNotFoundError(f"RESUME_EXCEL_PATH={env_path} does not exist.")
    matches = list(Path(".").glob("*.xlsx"))
    if matches:
        return matches[0]
    raise FileNotFoundError(
        "No Excel file found. Set RESUME_EXCEL_PATH or place a .xlsx in the working directory."
    )


def _sheet_to_text(ws) -> str:
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            rows.append("\t".join("" if v is None else str(v) for v in row))
    return "\n".join(rows)


def _normalize(obj):
    """Recursively normalize Pydantic models and dicts."""
    if hasattr(obj, "model_dump"):
        return obj.model_dump()
    if isinstance(obj, dict):
        return {k: _normalize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_normalize(i) for i in obj]
    return obj


def _fmt_date(raw: str | None) -> str:
    if not raw:
        return ""
    raw = str(raw).strip()

    # Already clean: "May 2026" or "May 2024 – June 2024"
    if re.match(r'^[A-Za-z]+ \d{4}', raw):
        return raw

    # Split on em dash or spaced hyphen ONLY — bare "-" appears inside ISO dates
    range_match = re.split(r'\s*–\s*|\s+-\s+', raw, maxsplit=1)
    if len(range_match) == 2:
        left  = _fmt_date(range_match[0].strip())
        right = _fmt_date(range_match[1].strip())
        if left and right:
            return f"{left} – {right}"

    # ISO / datetime strings
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw[:19], fmt).strftime("%B %Y")
        except ValueError:
            continue

    # Excel numeric serial date
    try:
        serial = float(raw)
        from datetime import timedelta
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(serial))).strftime("%B %Y")
    except (ValueError, TypeError):
        pass

    return raw


def _fmt_gpa(raw: str | None) -> str:
    """
    Extract just the numeric GPA from strings like '3.9/4.0 (exceeds 3.5 requirement)'.
    Returns the first decimal number found, e.g. '3.9'.
    """
    if not raw:
        return ""
    match = re.search(r'\d+\.\d+', str(raw))
    return match.group(0) if match else str(raw).strip()


# ---------------------------------------------------------------------------
# Excel tools
# ---------------------------------------------------------------------------

def load_all_excel_data() -> str:
    """Reads all sheets from the resume Excel workbook and returns them concatenated."""
    wb = openpyxl.load_workbook(_find_workbook(), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== SHEET: {sheet_name.upper()} ===\n{_sheet_to_text(ws)}")
    return "\n\n".join(parts)


# ---------------------------------------------------------------------------
# PDF renderer — matches Sarah's resume format
# ---------------------------------------------------------------------------

def save_resume_as_pdf(resume_bundle: dict | str, job_title: str = "role") -> str:
    """
    Renders the ResumeBundle as a PDF matching Sarah's resume format:
    Name header → contact line → Education → Work Experience → Projects → Skills
    """
    # Normalize input
    if isinstance(resume_bundle, str):
        try:
            resume_bundle = json.loads(resume_bundle)
        except Exception:
            resume_bundle = json.loads(
                resume_bundle.encode("utf-8").decode("unicode_escape")
            )
    resume_bundle = _normalize(resume_bundle)
    if not isinstance(resume_bundle, dict):
        raise ValueError(f"Could not parse resume_bundle, got type: {type(resume_bundle)}")

    # Pull fields
    personal    = resume_bundle.get("personal_info") or {}
    education   = resume_bundle.get("education") or []
    work_exp    = resume_bundle.get("work_experience") or []
    projects    = resume_bundle.get("projects") or []
    tech_skills = resume_bundle.get("technical_skills") or []

    def _build(fs: int) -> bytes:
        buf = BytesIO()
        lead = fs + 2

        doc = SimpleDocTemplate(
            buf, pagesize=letter,
            leftMargin=0.65*inch, rightMargin=0.65*inch,
            topMargin=0.5*inch,  bottomMargin=0.5*inch,
        )

        black = colors.HexColor("#000000")
        gray  = colors.HexColor("#444444")

        # ── Styles ────────────────────────────────────────────────────────
        name_style = ParagraphStyle(
            "Name",
            fontSize=fs + 8,
            fontName="Helvetica-Bold",
            spaceBefore=0,
            spaceAfter=6,       # breathing room between name and contact line
            textColor=black,
        )

        contact_style = ParagraphStyle(
            "Contact",
            fontSize=fs,        # same size as body (was fs-1, too small)
            fontName="Helvetica",
            spaceBefore=0,
            spaceAfter=8,       # gap before the rule
            textColor=gray,
            leading=fs + 4,
        )

        section_style = ParagraphStyle(
            "Section",
            fontSize=fs + 1,
            fontName="Helvetica-Bold",
            spaceBefore=6,
            spaceAfter=1,
            textColor=black,
        )

        body_style = ParagraphStyle(
            "Body",
            fontSize=fs,
            leading=lead,
            spaceAfter=1,
            textColor=black,
        )

        # Right-aligned style used exclusively for date cells
        date_style = ParagraphStyle(
            "Date",
            fontSize=fs,
            leading=lead,
            spaceAfter=1,
            textColor=black,
            alignment=TA_RIGHT,
        )

        role_style = ParagraphStyle(
            "Role",
            fontSize=fs,
            fontName="Helvetica-Bold",
            spaceAfter=0,
            textColor=black,
        )

        org_style = ParagraphStyle(
            "Org",
            fontSize=fs,
            fontName="Helvetica-Oblique",
            spaceAfter=1,
            textColor=black,
        )

        bullet_style = ParagraphStyle(
            "Bullet",
            fontSize=fs,
            leading=lead,
            leftIndent=10,
            firstLineIndent=-10,
            spaceAfter=1,
            textColor=black,
        )

        def section_header(title):
            return [
                Paragraph(title, section_style),
                HRFlowable(width="100%", thickness=0.5, color=black, spaceAfter=2),
            ]

        def two_col(left_para, right_text):
            """Left content with right-aligned date. Uses date_style for right cell."""
            tbl = Table(
                [[left_para, Paragraph(right_text, date_style)]],
                colWidths=["75%", "25%"],
            )
            tbl.setStyle(TableStyle([
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ("ALIGN",         (1, 0), (1,  0),  "RIGHT"),
                ("LEFTPADDING",   (0, 0), (-1, -1), 0),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
                ("TOPPADDING",    (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            return tbl

        story = []

        # ── Header ───────────────────────────────────────────────────────
        name = personal.get("name") or "Candidate"
        story.append(Paragraph(name, name_style))

        contact_parts = []
        if personal.get("phone"):
            contact_parts.append(personal["phone"])
        if personal.get("location"):
            contact_parts.append(personal["location"])
        if personal.get("email"):
            contact_parts.append(personal["email"])
        if personal.get("linkedin"):
            contact_parts.append(personal["linkedin"])
        if contact_parts:
            story.append(Paragraph("  |  ".join(contact_parts), contact_style))

        story.append(HRFlowable(width="100%", thickness=1, color=black, spaceAfter=6))

        # ── Education ────────────────────────────────────────────────────
        if education:
            story += section_header("Education")
            for edu in education:
                school = edu.get("school") or ""
                degree = edu.get("degree") or ""
                gpa    = _fmt_gpa(edu.get("gpa"))
                date   = _fmt_date(edu.get("date"))

                gpa_str = f" | GPA: {gpa}" if gpa else ""

                # School (bold) on first line, degree + GPA on second line —
                # all in the left cell so the date sits cleanly on the right.
                left_content = f"<b>{school}</b><br/>{degree}{gpa_str}"
                left = Paragraph(left_content, body_style)
                story.append(two_col(left, date))
                story.append(Spacer(1, 4))

        # ── Work Experience ──────────────────────────────────────────────
        if work_exp:
            story += section_header("Work Experience")
            for job in work_exp:
                role    = job.get("role") or ""
                company = job.get("company") or ""
                dates   = _fmt_date(job.get("date_range") or "")
                bullets = job.get("bullets") or []

                left = Paragraph(f"<b>{role}</b>", body_style)
                story.append(two_col(left, dates))
                story.append(Paragraph(company, org_style))
                for b in bullets:
                    story.append(Paragraph(f"- {b}", bullet_style))
                story.append(Spacer(1, 2))

        # ── Project Experience ───────────────────────────────────────────
        if projects:
            story += section_header("Project Experience")
            for proj in projects:
                title   = proj.get("title") or ""
                bullets = proj.get("bullets") or []
                story.append(Paragraph(f"<b>{title}</b>", role_style))
                for b in bullets:
                    story.append(Paragraph(f"- {b}", bullet_style))
                story.append(Spacer(1, 2))

        # ── Technical Skills ─────────────────────────────────────────────
        if tech_skills:
            story += section_header("Technical Skills")
            for cat in tech_skills:
                category = cat.get("category") or ""
                skills   = cat.get("skills") or ""
                story.append(Paragraph(f"<b>{category}:</b> {skills}", body_style))

        doc.build(story)
        return buf.getvalue()

    # Shrink font until single page (try down to 8pt)
    pdf_bytes = _build(10)
    for fs in range(10, 7, -1):
        pdf_bytes = _build(fs)
        if len(PdfReader(BytesIO(pdf_bytes)).pages) == 1:
            break

    # If still over 1 page at 8pt, trim oldest project and retry
    if len(PdfReader(BytesIO(pdf_bytes)).pages) > 1:
        projects = projects[:-1]
        pdf_bytes = _build(9)

    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in job_title).strip("_")[:60]
    pdf_path = output_dir / f"resume_{safe}.pdf"
    pdf_path.write_bytes(pdf_bytes)

    print(f"[save_resume_as_pdf] saved → {pdf_path}")
    return str(pdf_path)